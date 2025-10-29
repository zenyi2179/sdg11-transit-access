"""
基尼系数计算模块

主要功能：
- 计算公共交通可达性基尼系数
- 评估服务公平性
- 生成公平性报告
"""

import numpy as np
import pandas as pd
import openpyxl
from typing import Dict, List, Tuple, Optional
import csv


class GiniCalculator:
    """公共交通可达性基尼系数计算器"""
    
    def __init__(self):
        """初始化基尼计算器"""
        np.set_printoptions(suppress=True)  # 取消科学计数法
    
    def calculate_basic_stats(self, data: List[List]) -> Dict:
        """
        计算基本统计信息
        
        Args:
            data: 输入数据列表
            
        Returns:
            统计信息字典
        """
        popu_all = popu_valid = total_gather = 0
        bus_count = railway_count = tram_count = 0
        
        for row in data:
            popu_all += float(row[2])  # 总人口
            if row[6] != '0':  # 有效服务人口
                popu_valid += float(row[2])
                total_gather += int(row[6])
                if row[3] != '0': 
                    bus_count += int(row[3])  # 公交站点数
                if row[4] != '0': 
                    railway_count += int(row[4])  # 火车站数
                if row[5] != '0': 
                    tram_count += int(row[5])  # 有轨电车站数
        
        # 计算比例
        popu_ratio = popu_valid / popu_all if popu_all > 0 else 0
        bus_ratio = bus_count / total_gather if total_gather > 0 else 0
        railway_ratio = railway_count / total_gather if total_gather > 0 else 0
        tram_ratio = tram_count / total_gather if total_gather > 0 else 0
        
        return {
            'total_population': round(popu_all, 4),
            'valid_population': round(popu_valid, 4),
            'population_ratio': round(popu_ratio, 4),
            'bus_ratio': round(bus_ratio, 4),
            'railway_ratio': round(railway_ratio, 4),
            'tram_ratio': round(tram_ratio, 4)
        }
    
    def prepare_gini_data(self, data: List[List]) -> np.ndarray:
        """
        准备基尼系数计算数据
        
        Args:
            data: 原始数据列表
            
        Returns:
            汇总数据表
        """
        gather_dict = {}  # 按服务等级分组
        
        for row in data:
            gather_level = int(row[6])  # 服务等级
            population = float(row[2])  # 人口数量
            
            if gather_level != 0:
                if gather_level not in gather_dict:
                    gather_dict[gather_level] = population
                else:
                    gather_dict[gather_level] += population
        
        # 转换为数组
        gather_levels = list(gather_dict.keys())
        populations = list(gather_dict.values())
        ids = list(range(1, len(gather_levels) + 1))
        
        # 计算加权值
        total_population = sum(populations)
        weighted_values = [pop * level for pop, level in zip(populations, gather_levels)]
        total_weighted = sum(weighted_values)
        
        # 计算各种比例
        pop_ratios = [pop / total_population for pop in populations]
        weight_ratios = [weight / total_weighted for weight in weighted_values]
        weight_pop_ratios = [w_ratio / p_ratio if p_ratio > 0 else 0 
                           for w_ratio, p_ratio in zip(weight_ratios, pop_ratios)]
        
        # 构建汇总表
        summary_table = np.column_stack([
            ids,                           # ID
            gather_levels,                 # 服务等级
            populations,                   # 人口数
            weighted_values,               # 加权人口
            pop_ratios,                    # 人口比例
            weight_ratios,                 # 加权比例
            weight_pop_ratios              # 加权/人口比例
        ])
        
        return summary_table
    
    def calculate_gini_coefficient(self, summary_table: np.ndarray) -> Tuple[float, Dict]:
        """
        计算基尼系数
        
        Args:
            summary_table: 汇总数据表
            
        Returns:
            (基尼系数, 计算结果字典)
        """
        # 按加权/人口比例排序
        sorted_table = summary_table[summary_table[:, 6].argsort()]
        
        # 计算累积比例
        cum_pop_ratio = [0]  # 累积人口比例
        cum_weight_ratio = [0]  # 累积加权比例
        
        for i in range(len(sorted_table)):
            cum_pop_ratio.append(sorted_table[i, 4] + cum_pop_ratio[i])
            cum_weight_ratio.append(sorted_table[i, 5] + cum_weight_ratio[i])
        
        # 计算基尼系数所需的中间值
        total_population = sum(sorted_table[:, 2])
        total_weighted = sum(sorted_table[:, 3])
        pw_product = total_population * total_weighted
        
        # 计算 Σ(Wi-1+Wi)*Pi
        cumulative_sum = 0
        calculations = []
        
        for i in range(len(sorted_table)):
            if i == 0:
                cumulative_weight = sorted_table[i, 3]
                calculation = cumulative_weight * sorted_table[i, 2]
            else:
                prev_cumulative = cumulative_weight
                cumulative_weight += sorted_table[i, 3]
                calculation = (prev_cumulative + cumulative_weight) * sorted_table[i, 2]
            
            cumulative_sum += calculation
            calculations.append(calculation)
        
        # 计算基尼系数
        gini_coefficient = 1 - (cumulative_sum / pw_product) if pw_product > 0 else 0
        
        results = {
            'gini_coefficient': round(gini_coefficient, 4),
            'cumulative_sum': round(cumulative_sum, 4),
            'pw_product': round(pw_product, 4),
            'cum_pop_ratio': cum_pop_ratio,
            'cum_weight_ratio': cum_weight_ratio,
            'sorted_table': sorted_table,
            'calculations': calculations
        }
        
        return gini_coefficient, results
    
    def generate_gini_report(self, input_file: str, output_file: str = None) -> Dict:
        """
        生成基尼系数分析报告
        
        Args:
            input_file: 输入数据文件路径
            output_file: 输出Excel文件路径
            
        Returns:
            分析结果字典
        """
        # 读取数据
        data = []
        with open(input_file, 'r', encoding='utf-8') as f:
            csv_reader = csv.reader(f)
            header = next(csv_reader)  # 跳过标题行
            for row in csv_reader:
                data.append(row)
        
        # 计算基本统计信息
        basic_stats = self.calculate_basic_stats(data)
        
        # 准备基尼计算数据
        summary_table = self.prepare_gini_data(data)
        
        # 计算基尼系数
        gini_coefficient, gini_results = self.calculate_gini_coefficient(summary_table)
        
        # 生成输出报告
        if output_file:
            self._export_to_excel(
                data, basic_stats, summary_table, gini_results, 
                gini_coefficient, output_file
            )
        
        # 返回完整结果
        results = {
            'basic_statistics': basic_stats,
            'gini_coefficient': gini_coefficient,
            'gini_details': gini_results
        }
        
        return results
    
    def _export_to_excel(self, original_data: List, basic_stats: Dict,
                        summary_table: np.ndarray, gini_results: Dict,
                        gini_coefficient: float, output_file: str):
        """
        导出结果到Excel文件
        
        Args:
            original_data: 原始数据
            basic_stats: 基本统计信息
            summary_table: 汇总表
            gini_results: 基尼计算结果
            gini_coefficient: 基尼系数
            output_file: 输出文件路径
        """
        wb = openpyxl.Workbook()
        
        # 原始数据表
        sheet_original = wb.active
        sheet_original.title = "原始数据"
        
        # 基尼计算表
        sheet_gini = wb.create_sheet("基尼计算")
        
        # 写入基尼计算数据
        headers = ['ID', '服务等级', '人口', '加权值', '人口比例', '加权比例', '加权/人口比例']
        for i, header in enumerate(headers, 1):
            sheet_gini.cell(1, i, header)
        
        for i, row in enumerate(summary_table, 2):
            for j, value in enumerate(row, 1):
                sheet_gini.cell(i, j, value)
        
        # 写入排序后数据
        start_row = len(summary_table) + 4
        for i, header in enumerate(headers, 1):
            sheet_gini.cell(start_row, i, header)
        
        sorted_table = gini_results['sorted_table']
        for i, row in enumerate(sorted_table, start_row + 1):
            for j, value in enumerate(row, 1):
                sheet_gini.cell(i, j, value)
        
        # 写入基尼系数结果
        result_row = start_row + len(sorted_table) + 3
        sheet_gini.cell(result_row, 1, "基尼系数计算结果")
        sheet_gini.cell(result_row + 1, 1, "基尼系数")
        sheet_gini.cell(result_row + 1, 2, gini_coefficient)
        
        wb.save(output_file)